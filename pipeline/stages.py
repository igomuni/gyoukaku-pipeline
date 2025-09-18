import os
import csv
import zipfile
import logging
from typing import Callable, Optional, List
from pathlib import Path

import pandas as pd
import openpyxl

from config import (
    DOWNLOAD_DIR, RAW_DIR, NORMALIZED_DIR, PROCESSED_DIR, FILENAME_YEAR_MAP
)
from utils.normalization import normalize_text

# --- 処理ロジックのインポート ---
from pipeline.business_processing import build_business_tables
from pipeline.budget_processing import process_budget_files, PAST_BUDGET_ITEMS, REQUEST_BUDGET_ITEMS
from pipeline.fund_flow_processing import process_fund_flow
from pipeline.expenditure_processing import process_expenditures, EXPENDITURE_LIST_ITEMS


# ロガーの設定
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# --- Stage 1: Convert Excel/ZIP to CSV ---
def run_stage_01_convert(update_status: Callable, job_id: str, target_files: Optional[List[str]]):
    update_status(current_stage="ステージ1: CSVへの変換", message="処理を開始します...")
    
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    
    source_paths = list(DOWNLOAD_DIR.glob('*.zip')) + list(DOWNLOAD_DIR.glob('*.xlsx'))

    if target_files:
        source_paths = [p for p in source_paths if p.name in target_files]

    if not source_paths:
        logging.warning("[Stage 1] No target files found. Skipping.")
        update_status(message="対象ファイルが見つかりません。スキップします。")
        return

    def _convert_excel_to_csv(excel_source, file_stem, output_dir):
        try:
            workbook = openpyxl.load_workbook(excel_source, read_only=True, data_only=True)
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                output_path = output_dir / f"{file_stem}_{sheet_name}.csv"
                logging.info(f"  - Saving sheet: '{sheet_name}' -> '{output_path.name}'")
                with open(output_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
                    csv_writer = csv.writer(csv_file, quoting=csv.QUOTE_ALL)
                    for row in worksheet.iter_rows(values_only=True):
                        escaped_row = [str(cell).replace('\n', '\\n').replace('\r', '') if cell is not None else "" for cell in row]
                        csv_writer.writerow(escaped_row)
        except Exception as e:
            logging.error(f"  [ERROR] Failed to process Excel data from {file_stem}: {e}", exc_info=True)
            raise

    total_files = len(source_paths)
    for i, path in enumerate(source_paths):
        update_status(message=f"ファイル {i+1}/{total_files} を処理中: {path.name}")
        logging.info(f"Processing '{path.name}'...")
        if path.suffix == '.zip':
            try:
                with zipfile.ZipFile(path, 'r') as zf:
                    for file_in_zip in zf.namelist():
                        if file_in_zip.endswith('.xlsx') and not file_in_zip.startswith('__MACOSX'):
                            file_stem = Path(file_in_zip).stem
                            logging.info(f"  - Extracting '{file_in_zip}'")
                            with zf.open(file_in_zip) as excel_stream:
                                _convert_excel_to_csv(excel_stream, file_stem, RAW_DIR)
            except Exception as e:
                logging.error(f"  [ERROR] Failed to process zip file {path.name}: {e}", exc_info=True)
                raise
        elif path.suffix == '.xlsx':
            _convert_excel_to_csv(path, path.stem, RAW_DIR)
    
    update_status(message="ステージ1が完了しました。")

# --- Stage 2: Normalize CSV Files ---
def run_stage_02_normalize(update_status: Callable, job_id: str):
    update_status(current_stage="ステージ2: データの正規化", message="処理を開始します...")

    NORMALIZED_DIR.mkdir(parents=True, exist_ok=True)

    csv_files = sorted(list(RAW_DIR.glob('*.csv')))
    if not csv_files:
        logging.warning("[Stage 2] No .csv files found in 'data/raw/'. Skipping.")
        update_status(message="対象ファイルが見つかりません。スキップします。")
        return

    total_files = len(csv_files)
    for i, input_path in enumerate(csv_files):
        update_status(message=f"ファイル {i+1}/{total_files} を正規化中: {input_path.name}")
        output_path = NORMALIZED_DIR / input_path.name
        try:
            with open(input_path, 'r', encoding='utf-8-sig') as infile, \
                 open(output_path, 'w', encoding='utf-8-sig', newline='') as outfile:
                reader = csv.reader(infile)
                writer = csv.writer(outfile, quoting=csv.QUOTE_ALL)
                
                header = next(reader, None)
                if header:
                    writer.writerow([normalize_text(cell) for cell in header])
                
                for row in reader:
                    normalized_row = [normalize_text(cell.replace('\\n', '\n')) for cell in row]
                    writer.writerow(normalized_row)

        except Exception as e:
            logging.error(f"  [ERROR] Failed to process {input_path.name}: {e}", exc_info=True)
            raise

    update_status(message="ステージ2が完了しました。")


# --- Stage 3: Build Business Tables ---
def run_stage_03_build_business_tables(update_status: Callable, job_id: str):
    """
    ステージ3: 事業テーブルの構築（外部モジュール呼び出し）
    """
    build_business_tables(update_status, job_id)

# --- Stage 4: Build Budget Summary ---
def run_stage_04_build_budget_summary(update_status: Callable, job_id: str):
    """
    ステージ4: 予算テーブルの構築
    """
    update_status(current_stage="ステージ4: 予算テーブルの構築", message="処理を開始します...")
    
    all_csv_files = sorted(list(NORMALIZED_DIR.glob('*.csv')))
    if not all_csv_files:
        logging.warning("[Stage 4] No normalized CSV files found. Skipping.")
        update_status(message="正規化済みCSVが見つかりません。スキップします。")
        return
        
    review_year_map = {f.stem: get_year_from_filename(f.name) for f in all_csv_files}
    final_df = process_budget_files(all_csv_files, review_year_map)

    if final_df.empty:
        logging.warning("[Stage 4] No budget data could be extracted.")
        update_status(message="抽出対象の予算データが見つかりませんでした。")
        return

    final_columns = ['business_id']
    for i in range(-3, 1):
        suffix = {-3: '_py3', -2: '_py2', -1: '_py1', 0: ''}[i]
        for item in PAST_BUDGET_ITEMS:
            final_columns.append(f"{item}{suffix}")
    for item in REQUEST_BUDGET_ITEMS:
        final_columns.append(f"{item}_req")
        
    ordered_cols = [col for col in final_columns if col in final_df.columns]
    other_cols = [col for col in final_df.columns if col not in ordered_cols]
    final_df = final_df[ordered_cols + other_cols]

    output_path = PROCESSED_DIR / "budgets.csv"
    final_df.to_csv(output_path, index=False, encoding='utf-8-sig')
    
    update_status(message=f"ステージ4が完了しました。{len(final_df)}件のデータを保存しました。")


# --- Stage 5: Build Fund Flow Table ---
def run_stage_05_build_fund_flow(update_status: Callable, job_id: str):
    """
    ステージ5: 資金の流れテーブルの構築
    """
    update_status(current_stage="ステージ5: 資金の流れテーブル構築", message="処理を開始します...")
    
    all_csv_files = sorted(list(NORMALIZED_DIR.glob('*.csv')))
    if not all_csv_files:
        logging.warning("[Stage 5] No normalized CSV files found. Skipping.")
        update_status(message="正規化済みCSVが見つかりません。スキップします。")
        return
    
    final_df = process_fund_flow(all_csv_files)
    
    if not final_df.empty:
        output_columns = [
            'business_id', 'block_id', 'sequence', 
            '支払先費目', '支払先使途', '支払先金額(百万円)', '支払先計'
        ]
        final_df = final_df.reindex(columns=output_columns)
        
        output_path = PROCESSED_DIR / "fund_flow.csv"
        final_df.to_csv(output_path, index=False, encoding='utf-8-sig')
        update_status(message=f"ステージ5が完了しました。{len(final_df)}件のデータを保存しました。")
    else:
        update_status(message="ステージ5は完了しましたが、対象データは見つかりませんでした。")


# --- Stage 6: Build Expenditure Table ---
def run_stage_06_build_expenditure(update_status: Callable, job_id: str):
    """
    ステージ6: 支出テーブルの構築
    """
    update_status(current_stage="ステージ6: 支出テーブル構築", message="処理を開始します...")

    all_csv_files = sorted(list(NORMALIZED_DIR.glob('*.csv')))
    if not all_csv_files:
        logging.warning("[Stage 6] No normalized CSV files found. Skipping.")
        update_status(message="正規化済みCSVが見つかりません。スキップします。")
        return
        
    final_df = process_expenditures(all_csv_files)
    
    if not final_df.empty:
        base_cols = ['business_id', 'block_id', 'sequence']
        output_columns = base_cols + EXPENDITURE_LIST_ITEMS
        final_df = final_df.reindex(columns=output_columns)
        
        output_path = PROCESSED_DIR / "expenditure.csv"
        final_df.to_csv(output_path, index=False, encoding='utf-8-sig')
        update_status(message=f"ステージ6が完了しました。{len(final_df)}件のデータを保存しました。")
    else:
        update_status(message="ステージ6は完了しましたが、対象データは見つかりませんでした。")


def get_year_from_filename(filename):
    """(stages.py内で使うためのヘルパー関数)"""
    for key, year in FILENAME_YEAR_MAP.items():
        if key in filename:
            return year
    return None